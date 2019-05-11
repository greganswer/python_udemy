import openpyxl
import os

os.chdir('/Users/greganswer/code/personal/practice/python_udemy')

workbook = openpyxl.load_workbook('example.xlsx')
workbook.sheetnames
sheet = workbook['Sheet1']
cell = sheet['A1']
cell.value
sheet.cell(row=1, column=2)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)

new_workbook = openpyxl.Workbook()
new_workbook_sheet1 = new_workbook['Sheet']

new_workbook_sheet1.title = 'Hello'
new_workbook_sheet1['A1'] = 'Hello'
new_workbook_sheet1['A2'] = 'World'

new_workbook_sheet2 = new_workbook.create_sheet(title="World", index=1)

new_workbook.save('new_example.xlsx')